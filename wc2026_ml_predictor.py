"""
WC2026 Group Stage Predictor — ML Edition
Trains a Ridge regression (Poisson-GLM style) on ~10 years of international
football to learn feature weights, then simulates the 2026 group stage.

Features per match (team1 perspective):
  1. ELO differential
  2. Height differential (set-piece proxy)
  3. Climate/heat penalty differential (age × temp-mismatch)
  4. Head-to-head historical advantage
  5. Opponent GK quality
  6. Form differential
  7. Underdog motivation boost
"""

import csv
import math
import sys
from collections import defaultdict

import numpy as np
from sklearn.linear_model import Ridge
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import cross_val_score

# ============================================================
# 2026 CURRENT TEAM DATA
# Sources: eloratings.net, transfermarkt (June 2026 approx)
# ============================================================
TEAMS = {
    "France":               {"elo": 2045, "height": 183.2, "age": 27.1, "gk": 9.2, "form": 0.72, "conf": "UEFA", "sp": 7.5},
    "Argentina":            {"elo": 2005, "height": 180.1, "age": 27.8, "gk": 9.0, "form": 0.68, "conf": "CONMEBOL", "sp": 6.5},
    "England":              {"elo": 1981, "height": 183.5, "age": 26.5, "gk": 8.8, "form": 0.70, "conf": "UEFA", "sp": 9.5},
    "Brazil":               {"elo": 1968, "height": 180.8, "age": 26.2, "gk": 8.5, "form": 0.65, "conf": "CONMEBOL", "sp": 4.5},
    "Spain":                {"elo": 1962, "height": 181.0, "age": 25.8, "gk": 8.6, "form": 0.72, "conf": "UEFA", "sp": 5.0},
    "Portugal":             {"elo": 1951, "height": 182.4, "age": 27.9, "gk": 8.4, "form": 0.68, "conf": "UEFA", "sp": 6.5},
    "Germany":              {"elo": 1940, "height": 183.8, "age": 25.6, "gk": 8.7, "form": 0.65, "conf": "UEFA", "sp": 8.5},
    "Netherlands":          {"elo": 1905, "height": 185.2, "age": 26.8, "gk": 8.3, "form": 0.67, "conf": "UEFA", "sp": 8.5},
    "Belgium":              {"elo": 1885, "height": 182.6, "age": 28.4, "gk": 8.1, "form": 0.62, "conf": "UEFA", "sp": 7.5},
    "Uruguay":              {"elo": 1858, "height": 181.3, "age": 27.0, "gk": 7.8, "form": 0.63, "conf": "CONMEBOL", "sp": 7.0},
    "Croatia":              {"elo": 1840, "height": 184.1, "age": 29.2, "gk": 8.5, "form": 0.58, "conf": "UEFA", "sp": 7.5},
    "United States":        {"elo": 1830, "height": 183.0, "age": 25.2, "gk": 7.6, "form": 0.62, "conf": "CONCACAF", "sp": 5.5},
    "Colombia":             {"elo": 1818, "height": 181.2, "age": 26.4, "gk": 7.8, "form": 0.66, "conf": "CONMEBOL", "sp": 5.5},
    "Switzerland":          {"elo": 1800, "height": 182.8, "age": 27.6, "gk": 8.2, "form": 0.64, "conf": "UEFA", "sp": 6.5},
    "Mexico":               {"elo": 1795, "height": 178.4, "age": 26.8, "gk": 7.5, "form": 0.58, "conf": "CONCACAF", "sp": 5.0},
    "Senegal":              {"elo": 1785, "height": 183.0, "age": 26.5, "gk": 8.0, "form": 0.65, "conf": "CAF", "sp": 7.0},
    "Japan":                {"elo": 1778, "height": 175.8, "age": 26.2, "gk": 7.9, "form": 0.68, "conf": "AFC", "sp": 4.0},
    "Morocco":              {"elo": 1770, "height": 181.5, "age": 26.8, "gk": 8.4, "form": 0.64, "conf": "CAF", "sp": 7.5},
    "Canada":               {"elo": 1752, "height": 183.6, "age": 25.8, "gk": 7.8, "form": 0.60, "conf": "CONCACAF", "sp": 6.0},
    "South Korea":          {"elo": 1738, "height": 180.2, "age": 27.1, "gk": 7.6, "form": 0.58, "conf": "AFC", "sp": 5.0},
    "Ecuador":              {"elo": 1725, "height": 179.8, "age": 25.5, "gk": 7.4, "form": 0.56, "conf": "CONMEBOL", "sp": 5.5},
    "Ivory Coast":          {"elo": 1712, "height": 181.8, "age": 27.4, "gk": 7.5, "form": 0.58, "conf": "CAF", "sp": 5.5},
    "Australia":            {"elo": 1705, "height": 182.4, "age": 27.0, "gk": 7.6, "form": 0.56, "conf": "AFC", "sp": 5.5},
    "Sweden":               {"elo": 1698, "height": 185.0, "age": 27.8, "gk": 7.8, "form": 0.60, "conf": "UEFA", "sp": 9.0},
    "Norway":               {"elo": 1690, "height": 184.2, "age": 26.0, "gk": 7.4, "form": 0.58, "conf": "UEFA", "sp": 8.5},
    "Algeria":              {"elo": 1675, "height": 181.4, "age": 27.2, "gk": 7.5, "form": 0.55, "conf": "CAF", "sp": 6.0},
    "Austria":              {"elo": 1665, "height": 183.0, "age": 26.5, "gk": 7.6, "form": 0.60, "conf": "UEFA", "sp": 6.5},
    "Turkiye":              {"elo": 1658, "height": 183.2, "age": 26.8, "gk": 7.7, "form": 0.58, "conf": "UEFA", "sp": 6.5},
    "Paraguay":             {"elo": 1645, "height": 179.5, "age": 26.2, "gk": 7.2, "form": 0.52, "conf": "CONMEBOL", "sp": 5.0},
    "Iran":                 {"elo": 1638, "height": 183.8, "age": 28.2, "gk": 7.8, "form": 0.56, "conf": "AFC", "sp": 8.0},
    "Tunisia":              {"elo": 1625, "height": 181.6, "age": 27.5, "gk": 7.3, "form": 0.54, "conf": "CAF", "sp": 5.5},
    "Egypt":                {"elo": 1615, "height": 181.2, "age": 27.8, "gk": 7.4, "form": 0.55, "conf": "CAF", "sp": 5.5},
    "Ghana":                {"elo": 1602, "height": 181.0, "age": 26.0, "gk": 7.0, "form": 0.50, "conf": "CAF", "sp": 5.5},
    "Bosnia and Herzegovina": {"elo": 1595, "height": 184.2, "age": 27.5, "gk": 7.2, "form": 0.52, "conf": "UEFA", "sp": 7.5},
    "Scotland":             {"elo": 1580, "height": 183.5, "age": 27.8, "gk": 7.4, "form": 0.54, "conf": "UEFA", "sp": 7.0},
    "Saudi Arabia":         {"elo": 1568, "height": 180.8, "age": 26.5, "gk": 7.0, "form": 0.50, "conf": "AFC", "sp": 5.0},
    "Panama":               {"elo": 1548, "height": 181.2, "age": 27.2, "gk": 7.0, "form": 0.50, "conf": "CONCACAF", "sp": 5.0},
    "Jordan":               {"elo": 1532, "height": 180.5, "age": 26.8, "gk": 6.8, "form": 0.48, "conf": "AFC", "sp": 4.5},
    "New Zealand":          {"elo": 1515, "height": 183.8, "age": 26.5, "gk": 6.8, "form": 0.46, "conf": "OFC", "sp": 5.5},
    "DR Congo":             {"elo": 1505, "height": 182.2, "age": 26.2, "gk": 6.9, "form": 0.48, "conf": "CAF", "sp": 6.0},
    "Uzbekistan":           {"elo": 1492, "height": 181.8, "age": 25.5, "gk": 6.7, "form": 0.50, "conf": "AFC", "sp": 5.0},
    "Cape Verde":           {"elo": 1478, "height": 181.0, "age": 26.8, "gk": 6.8, "form": 0.52, "conf": "CAF", "sp": 5.5},
    "Czechia":              {"elo": 1748, "height": 184.0, "age": 27.8, "gk": 7.9, "form": 0.58, "conf": "UEFA", "sp": 7.0},
    "Iraq":                 {"elo": 1420, "height": 180.2, "age": 26.5, "gk": 6.6, "form": 0.46, "conf": "AFC", "sp": 5.0},
    "South Africa":         {"elo": 1412, "height": 181.5, "age": 26.8, "gk": 6.8, "form": 0.48, "conf": "CAF", "sp": 5.0},
    "Curacao":              {"elo": 1395, "height": 181.0, "age": 27.5, "gk": 6.5, "form": 0.44, "conf": "CONCACAF", "sp": 4.5},
    "Qatar":                {"elo": 1380, "height": 179.0, "age": 26.2, "gk": 6.4, "form": 0.40, "conf": "AFC", "sp": 4.0},
    "Haiti":                {"elo": 1355, "height": 180.5, "age": 26.5, "gk": 6.2, "form": 0.38, "conf": "CONCACAF", "sp": 4.5},
}

# ============================================================
# HISTORICAL TEAM DATA — used for training
# (team_name): approximate values at tournament time
# Heights are stable; ELO/form/age/GK approximated per year
# ============================================================

# Historical team ELO snapshots at key tournaments
HIST_ELO = {
    # --- WC 2022 (Nov-Dec 2022) ---
    ("WC2022", "Netherlands"):  1921,
    ("WC2022", "Senegal"):      1770,
    ("WC2022", "Ecuador"):      1710,
    ("WC2022", "Qatar"):        1595,
    ("WC2022", "England"):      1960,
    ("WC2022", "Iran"):         1720,
    ("WC2022", "USA"):          1780,
    ("WC2022", "Wales"):        1735,
    ("WC2022", "Argentina"):    1975,
    ("WC2022", "Saudi Arabia"): 1633,
    ("WC2022", "Mexico"):       1792,
    ("WC2022", "Poland"):       1757,
    ("WC2022", "France"):       2016,
    ("WC2022", "Australia"):    1698,
    ("WC2022", "Denmark"):      1870,
    ("WC2022", "Tunisia"):      1672,
    ("WC2022", "Spain"):        1966,
    ("WC2022", "Germany"):      1942,
    ("WC2022", "Japan"):        1754,
    ("WC2022", "Costa Rica"):   1662,
    ("WC2022", "Belgium"):      1906,
    ("WC2022", "Morocco"):      1760,
    ("WC2022", "Croatia"):      1845,
    ("WC2022", "Canada"):       1738,
    ("WC2022", "Brazil"):       2002,
    ("WC2022", "Cameroon"):     1645,
    ("WC2022", "Serbia"):       1797,
    ("WC2022", "Switzerland"):  1807,
    ("WC2022", "Portugal"):     1955,
    ("WC2022", "Uruguay"):      1845,
    ("WC2022", "South Korea"):  1723,
    ("WC2022", "Ghana"):        1621,
    # --- WC 2018 (June-July 2018) ---
    ("WC2018", "Russia"):       1685,
    ("WC2018", "Saudi Arabia"): 1582,
    ("WC2018", "Egypt"):        1646,
    ("WC2018", "Uruguay"):      1872,
    ("WC2018", "Portugal"):     1929,
    ("WC2018", "Spain"):        1960,
    ("WC2018", "Morocco"):      1715,
    ("WC2018", "Iran"):         1706,
    ("WC2018", "France"):       1984,
    ("WC2018", "Peru"):         1779,
    ("WC2018", "Denmark"):      1826,
    ("WC2018", "Australia"):    1682,
    ("WC2018", "Argentina"):    1979,
    ("WC2018", "Croatia"):      1834,
    ("WC2018", "Nigeria"):      1680,
    ("WC2018", "Iceland"):      1736,
    ("WC2018", "Brazil"):       2006,
    ("WC2018", "Switzerland"):  1879,
    ("WC2018", "Serbia"):       1782,
    ("WC2018", "Costa Rica"):   1688,
    ("WC2018", "Germany"):      2086,
    ("WC2018", "Mexico"):       1840,
    ("WC2018", "Sweden"):       1808,
    ("WC2018", "South Korea"):  1737,
    ("WC2018", "Belgium"):      1931,
    ("WC2018", "England"):      1902,
    ("WC2018", "Tunisia"):      1669,
    ("WC2018", "Panama"):       1656,
    ("WC2018", "Colombia"):     1876,
    ("WC2018", "Japan"):        1751,
    ("WC2018", "Senegal"):      1731,
    ("WC2018", "Poland"):       1817,
    # --- Euro 2020/2021 (June-July 2021, avg 20C) ---
    ("EURO2020", "Italy"):      1978,
    ("EURO2020", "Wales"):      1737,
    ("EURO2020", "Switzerland"):1808,
    ("EURO2020", "Turkey"):     1692,
    ("EURO2020", "Belgium"):    1920,
    ("EURO2020", "Finland"):    1682,
    ("EURO2020", "Denmark"):    1834,
    ("EURO2020", "Russia"):     1708,
    ("EURO2020", "Netherlands"):1905,
    ("EURO2020", "Austria"):    1733,
    ("EURO2020", "Ukraine"):    1724,
    ("EURO2020", "North Macedonia"): 1597,
    ("EURO2020", "England"):    1921,
    ("EURO2020", "Croatia"):    1841,
    ("EURO2020", "Czech Republic"): 1762,
    ("EURO2020", "Scotland"):   1699,
    ("EURO2020", "Spain"):      1952,
    ("EURO2020", "Sweden"):     1787,
    ("EURO2020", "Poland"):     1805,
    ("EURO2020", "Slovakia"):   1712,
    ("EURO2020", "Hungary"):    1699,
    ("EURO2020", "France"):     2012,
    ("EURO2020", "Germany"):    1988,
    ("EURO2020", "Portugal"):   1946,
    # --- Euro 2024 (June 2024, Germany, avg 22C) ---
    ("EURO2024", "Germany"):    1951,
    ("EURO2024", "Scotland"):   1679,
    ("EURO2024", "Hungary"):    1724,
    ("EURO2024", "Switzerland"):1873,
    ("EURO2024", "Spain"):      1971,
    ("EURO2024", "Italy"):      1937,
    ("EURO2024", "Croatia"):    1831,
    ("EURO2024", "Albania"):    1649,
    ("EURO2024", "Slovenia"):   1706,
    ("EURO2024", "Serbia"):     1744,
    ("EURO2024", "England"):    1972,
    ("EURO2024", "Denmark"):    1841,
    ("EURO2024", "Poland"):     1770,
    ("EURO2024", "Netherlands"):1905,
    ("EURO2024", "Austria"):    1776,
    ("EURO2024", "France"):     2024,
    ("EURO2024", "Turkey"):     1756,
    ("EURO2024", "Georgia"):    1643,
    ("EURO2024", "Portugal"):   1961,
    ("EURO2024", "Czech Republic"): 1768,
    ("EURO2024", "Belgium"):    1893,
    ("EURO2024", "Romania"):    1701,
    ("EURO2024", "Slovakia"):   1722,
    ("EURO2024", "Ukraine"):    1731,
    # --- Copa América 2021 (June-July 2021, avg 22C, Brazil) ---
    ("COPA2021", "Brazil"):     2003,
    ("COPA2021", "Venezuela"):  1626,
    ("COPA2021", "Colombia"):   1826,
    ("COPA2021", "Ecuador"):    1716,
    ("COPA2021", "Peru"):       1759,
    ("COPA2021", "Argentina"):  1972,
    ("COPA2021", "Uruguay"):    1847,
    ("COPA2021", "Chile"):      1836,
    ("COPA2021", "Paraguay"):   1712,
    ("COPA2021", "Bolivia"):    1581,
    # --- Copa América 2024 (June-July 2024, USA, avg 27C) ---
    ("COPA2024", "Argentina"):  1995,
    ("COPA2024", "Canada"):     1762,
    ("COPA2024", "Chile"):      1813,
    ("COPA2024", "Peru"):       1729,
    ("COPA2024", "Colombia"):   1852,
    ("COPA2024", "Paraguay"):   1686,
    ("COPA2024", "Brazil"):     1990,
    ("COPA2024", "Costa Rica"): 1666,
    ("COPA2024", "Uruguay"):    1855,
    ("COPA2024", "Bolivia"):    1570,
    ("COPA2024", "Ecuador"):    1712,
    ("COPA2024", "Mexico"):     1808,
    ("COPA2024", "Venezuela"):  1670,
    ("COPA2024", "Jamaica"):    1572,
    ("COPA2024", "Panama"):     1660,
    ("COPA2024", "USA"):        1805,
    # --- AFCON 2021 (Jan-Feb 2022, Cameroon, avg 28C) ---
    ("AFCON2021", "Cameroon"):  1653,
    ("AFCON2021", "Burkina Faso"): 1632,
    ("AFCON2021", "Ethiopia"):  1432,
    ("AFCON2021", "Cape Verde"): 1545,
    ("AFCON2021", "Senegal"):   1769,
    ("AFCON2021", "Zimbabwe"):  1443,
    ("AFCON2021", "Guinea"):    1598,
    ("AFCON2021", "Malawi"):    1421,
    ("AFCON2021", "Morocco"):   1756,
    ("AFCON2021", "Ghana"):     1648,
    ("AFCON2021", "Comoros"):   1376,
    ("AFCON2021", "Gabon"):     1555,
    ("AFCON2021", "Tunisia"):   1694,
    ("AFCON2021", "Mali"):      1612,
    ("AFCON2021", "Mauritania"): 1430,
    ("AFCON2021", "Gambia"):    1492,
    ("AFCON2021", "Nigeria"):   1713,
    ("AFCON2021", "Egypt"):     1657,
    ("AFCON2021", "Sudan"):     1429,
    ("AFCON2021", "Guinea-Bissau"): 1479,
    ("AFCON2021", "Algeria"):   1767,
    ("AFCON2021", "Ivory Coast"): 1718,
    ("AFCON2021", "Sierra Leone"): 1416,
    ("AFCON2021", "Equatorial Guinea"): 1456,
    # --- WC 2014 (June-July 2014, Brazil, avg 27C) ---
    ("WC2014", "Brazil"):       1995,
    ("WC2014", "Croatia"):      1808,
    ("WC2014", "Mexico"):       1858,
    ("WC2014", "Cameroon"):     1629,
    ("WC2014", "Spain"):        2010,
    ("WC2014", "Netherlands"):  1913,
    ("WC2014", "Chile"):        1819,
    ("WC2014", "Australia"):    1650,
    ("WC2014", "Colombia"):     1840,
    ("WC2014", "Greece"):       1784,
    ("WC2014", "Ivory Coast"):  1726,
    ("WC2014", "Japan"):        1740,
    ("WC2014", "Uruguay"):      1882,
    ("WC2014", "Costa Rica"):   1666,
    ("WC2014", "England"):      1898,
    ("WC2014", "Italy"):        1927,
    ("WC2014", "Switzerland"):  1861,
    ("WC2014", "Ecuador"):      1694,
    ("WC2014", "France"):       1976,
    ("WC2014", "Honduras"):     1623,
    ("WC2014", "Argentina"):    1983,
    ("WC2014", "Bosnia"):       1748,
    ("WC2014", "Iran"):         1686,
    ("WC2014", "Nigeria"):      1693,
    ("WC2014", "Germany"):      2014,
    ("WC2014", "Portugal"):     1935,
    ("WC2014", "Ghana"):        1694,
    ("WC2014", "USA"):          1795,
    ("WC2014", "Belgium"):      1856,
    ("WC2014", "Algeria"):      1706,
    ("WC2014", "Russia"):       1785,
    ("WC2014", "South Korea"):  1738,
}

# Per-team stable features (height, home temp, confederation)
TEAM_META = {
    # team: {height_cm, home_temp_c, conf}
    "Netherlands":      {"height": 185.2, "home_temp": 12, "conf": "UEFA", "sp": 8.5},
    "Senegal":          {"height": 183.0, "home_temp": 30, "conf": "CAF", "sp": 7.0},
    "Ecuador":          {"height": 179.8, "home_temp": 18, "conf": "CONMEBOL", "sp": 5.5},
    "Qatar":            {"height": 179.0, "home_temp": 38, "conf": "AFC", "sp": 4.0},
    "England":          {"height": 183.5, "home_temp": 10, "conf": "UEFA", "sp": 9.5},
    "Iran":             {"height": 183.8, "home_temp": 28, "conf": "AFC", "sp": 8.0},
    "USA":              {"height": 183.0, "home_temp": 20, "conf": "CONCACAF"},
    "United States":    {"height": 183.0, "home_temp": 20, "conf": "CONCACAF", "sp": 5.5},
    "Wales":            {"height": 182.0, "home_temp": 10, "conf": "UEFA"},
    "Argentina":        {"height": 180.1, "home_temp": 15, "conf": "CONMEBOL", "sp": 6.5},
    "Saudi Arabia":     {"height": 180.8, "home_temp": 38, "conf": "AFC", "sp": 5.0},
    "Mexico":           {"height": 178.4, "home_temp": 22, "conf": "CONCACAF", "sp": 5.0},
    "Poland":           {"height": 184.0, "home_temp": 10, "conf": "UEFA"},
    "France":           {"height": 183.2, "home_temp": 12, "conf": "UEFA", "sp": 7.5},
    "Australia":        {"height": 182.4, "home_temp": 18, "conf": "AFC", "sp": 5.5},
    "Denmark":          {"height": 185.0, "home_temp": 10, "conf": "UEFA"},
    "Tunisia":          {"height": 181.6, "home_temp": 22, "conf": "CAF", "sp": 5.5},
    "Spain":            {"height": 181.0, "home_temp": 15, "conf": "UEFA", "sp": 5.0},
    "Germany":          {"height": 183.8, "home_temp": 12, "conf": "UEFA", "sp": 8.5},
    "Japan":            {"height": 175.8, "home_temp": 20, "conf": "AFC", "sp": 4.0},
    "Costa Rica":       {"height": 179.5, "home_temp": 25, "conf": "CONCACAF"},
    "Belgium":          {"height": 182.6, "home_temp": 10, "conf": "UEFA", "sp": 7.5},
    "Morocco":          {"height": 181.5, "home_temp": 22, "conf": "CAF", "sp": 7.5},
    "Croatia":          {"height": 184.1, "home_temp": 14, "conf": "UEFA", "sp": 7.5},
    "Canada":           {"height": 183.6, "home_temp": 10, "conf": "CONCACAF", "sp": 6.0},
    "Brazil":           {"height": 180.8, "home_temp": 26, "conf": "CONMEBOL", "sp": 4.5},
    "Cameroon":         {"height": 182.5, "home_temp": 28, "conf": "CAF"},
    "Serbia":           {"height": 184.5, "home_temp": 12, "conf": "UEFA"},
    "Switzerland":      {"height": 182.8, "home_temp": 10, "conf": "UEFA", "sp": 6.5},
    "Portugal":         {"height": 182.4, "home_temp": 16, "conf": "UEFA", "sp": 6.5},
    "Uruguay":          {"height": 181.3, "home_temp": 15, "conf": "CONMEBOL", "sp": 7.0},
    "South Korea":      {"height": 180.2, "home_temp": 18, "conf": "AFC", "sp": 5.0},
    "Ghana":            {"height": 181.0, "home_temp": 29, "conf": "CAF", "sp": 5.5},
    "Russia":           {"height": 183.0, "home_temp": 8,  "conf": "UEFA"},
    "Egypt":            {"height": 181.2, "home_temp": 30, "conf": "CAF", "sp": 5.5},
    "Peru":             {"height": 175.5, "home_temp": 18, "conf": "CONMEBOL"},
    "Iceland":          {"height": 184.0, "home_temp": 5,  "conf": "UEFA"},
    "Nigeria":          {"height": 181.5, "home_temp": 30, "conf": "CAF"},
    "Italy":            {"height": 183.5, "home_temp": 16, "conf": "UEFA"},
    "Turkey":           {"height": 183.2, "home_temp": 18, "conf": "UEFA"},
    "Turkiye":          {"height": 183.2, "home_temp": 18, "conf": "UEFA", "sp": 6.5},
    "Finland":          {"height": 182.5, "home_temp": 6,  "conf": "UEFA"},
    "North Macedonia":  {"height": 182.0, "home_temp": 14, "conf": "UEFA"},
    "Austria":          {"height": 183.0, "home_temp": 10, "conf": "UEFA", "sp": 6.5},
    "Ukraine":          {"height": 183.0, "home_temp": 10, "conf": "UEFA"},
    "Czech Republic":   {"height": 184.0, "home_temp": 10, "conf": "UEFA"},
    "Czechia":          {"height": 184.0, "home_temp": 10, "conf": "UEFA", "sp": 7.0},
    "Scotland":         {"height": 183.5, "home_temp": 8,  "conf": "UEFA", "sp": 7.0},
    "Sweden":           {"height": 185.0, "home_temp": 8,  "conf": "UEFA", "sp": 9.0},
    "Slovakia":         {"height": 183.5, "home_temp": 10, "conf": "UEFA"},
    "Hungary":          {"height": 182.0, "home_temp": 12, "conf": "UEFA"},
    "Colombia":         {"height": 181.2, "home_temp": 22, "conf": "CONMEBOL", "sp": 5.5},
    "Chile":            {"height": 178.5, "home_temp": 14, "conf": "CONMEBOL"},
    "Paraguay":         {"height": 179.5, "home_temp": 25, "conf": "CONMEBOL", "sp": 5.0},
    "Bolivia":          {"height": 173.0, "home_temp": 10, "conf": "CONMEBOL"},
    "Venezuela":        {"height": 179.5, "home_temp": 26, "conf": "CONMEBOL"},
    "Georgia":          {"height": 181.0, "home_temp": 14, "conf": "UEFA"},
    "Albania":          {"height": 181.5, "home_temp": 15, "conf": "UEFA"},
    "Slovenia":         {"height": 182.5, "home_temp": 11, "conf": "UEFA"},
    "Romania":          {"height": 182.0, "home_temp": 11, "conf": "UEFA"},
    "Ivory Coast":      {"height": 181.8, "home_temp": 30, "conf": "CAF", "sp": 5.5},
    "Algeria":          {"height": 181.4, "home_temp": 25, "conf": "CAF", "sp": 6.0},
    "Burkina Faso":     {"height": 178.0, "home_temp": 30, "conf": "CAF"},
    "Ethiopia":         {"height": 172.0, "home_temp": 22, "conf": "CAF"},
    "Cape Verde":       {"height": 181.0, "home_temp": 25, "conf": "CAF", "sp": 5.5},
    "Zimbabwe":         {"height": 179.5, "home_temp": 22, "conf": "CAF"},
    "Guinea":           {"height": 179.0, "home_temp": 28, "conf": "CAF"},
    "Malawi":           {"height": 176.0, "home_temp": 22, "conf": "CAF"},
    "Comoros":          {"height": 176.0, "home_temp": 28, "conf": "CAF"},
    "Gabon":            {"height": 180.0, "home_temp": 27, "conf": "CAF"},
    "Mali":             {"height": 180.0, "home_temp": 32, "conf": "CAF"},
    "Mauritania":       {"height": 178.0, "home_temp": 30, "conf": "CAF"},
    "Gambia":           {"height": 178.0, "home_temp": 28, "conf": "CAF"},
    "Nigeria":          {"height": 181.5, "home_temp": 30, "conf": "CAF"},
    "Sudan":            {"height": 177.0, "home_temp": 34, "conf": "CAF"},
    "Guinea-Bissau":    {"height": 178.0, "home_temp": 28, "conf": "CAF"},
    "Sierra Leone":     {"height": 178.0, "home_temp": 28, "conf": "CAF"},
    "Equatorial Guinea":{"height": 177.0, "home_temp": 28, "conf": "CAF"},
    "Jamaica":          {"height": 180.5, "home_temp": 28, "conf": "CONCACAF"},
    "Panama":           {"height": 181.2, "home_temp": 28, "conf": "CONCACAF", "sp": 5.0},
    "Greece":           {"height": 183.5, "home_temp": 16, "conf": "UEFA"},
    "Honduras":         {"height": 178.5, "home_temp": 26, "conf": "CONCACAF"},
    "Bosnia":           {"height": 184.2, "home_temp": 13, "conf": "UEFA"},
    "Bosnia and Herzegovina": {"height": 184.2, "home_temp": 13, "conf": "UEFA", "sp": 7.5},
    "Norway":           {"height": 184.2, "home_temp": 6,  "conf": "UEFA", "sp": 8.5},
    "Jordan":           {"height": 180.5, "home_temp": 28, "conf": "AFC", "sp": 4.5},
    "New Zealand":      {"height": 183.8, "home_temp": 14, "conf": "OFC", "sp": 5.5},
    "DR Congo":         {"height": 182.2, "home_temp": 28, "conf": "CAF", "sp": 6.0},
    "Uzbekistan":       {"height": 181.8, "home_temp": 25, "conf": "AFC", "sp": 5.0},
    "Curacao":          {"height": 181.0, "home_temp": 30, "conf": "CONCACAF", "sp": 4.5},
    "Haiti":            {"height": 180.5, "home_temp": 30, "conf": "CONCACAF", "sp": 4.5},
    "South Africa":     {"height": 181.5, "home_temp": 18, "conf": "CAF", "sp": 5.0},
    "Iraq":             {"height": 180.2, "home_temp": 40, "conf": "AFC", "sp": 5.0},
}

# GK quality approximations at tournament time (1-10 scale)
HIST_GK = {
    ("WC2022", "Netherlands"):  8.3,
    ("WC2022", "France"):       9.5,
    ("WC2022", "Brazil"):       8.5,
    ("WC2022", "Argentina"):    8.8,
    ("WC2022", "England"):      8.6,
    ("WC2022", "Spain"):        8.5,
    ("WC2022", "Germany"):      8.5,
    ("WC2022", "Portugal"):     8.2,
    ("WC2022", "Belgium"):      8.0,
    ("WC2022", "Morocco"):      8.6,
    ("WC2022", "Croatia"):      8.5,
    ("WC2022", "Switzerland"):  8.4,
    ("WC2022", "Senegal"):      8.2,
    ("WC2022", "Japan"):        8.0,
    ("WC2022", "South Korea"):  7.8,
    ("WC2022", "USA"):          7.5,
    ("WC2022", "Denmark"):      8.0,
    ("WC2022", "Poland"):       8.0,
    ("WC2022", "Uruguay"):      7.8,
    ("WC2022", "Australia"):    7.5,
    ("WC2022", "Ecuador"):      7.5,
    ("WC2022", "Canada"):       7.8,
    ("WC2022", "Serbia"):       7.8,
    ("WC2022", "Tunisia"):      7.4,
    ("WC2022", "Ghana"):        7.0,
    ("WC2022", "Cameroon"):     7.2,
    ("WC2022", "Iran"):         7.8,
    ("WC2022", "Mexico"):       7.5,
    ("WC2022", "Wales"):        7.6,
    ("WC2022", "Saudi Arabia"): 7.0,
    ("WC2022", "Costa Rica"):   7.6,
    ("WC2022", "Qatar"):        6.4,
    ("WC2018", "Germany"):      9.2,
    ("WC2018", "Spain"):        9.0,
    ("WC2018", "France"):       9.4,
    ("WC2018", "Brazil"):       8.4,
    ("WC2018", "Argentina"):    8.8,
    ("WC2018", "Belgium"):      8.2,
    ("WC2018", "Portugal"):     8.4,
    ("WC2018", "Croatia"):      8.5,
    ("WC2018", "England"):      8.2,
    ("WC2018", "Uruguay"):      7.8,
    ("WC2018", "Colombia"):     7.8,
    ("WC2018", "Switzerland"):  8.2,
    ("WC2018", "Denmark"):      8.0,
    ("WC2018", "Mexico"):       7.8,
    ("WC2018", "Sweden"):       7.8,
    ("WC2018", "Senegal"):      7.4,
    ("WC2018", "Japan"):        7.6,
    ("WC2018", "Russia"):       7.5,
    ("WC2018", "Poland"):       7.8,
    ("WC2018", "Peru"):         7.4,
    ("WC2018", "Iceland"):      8.0,
    ("WC2018", "Serbia"):       7.6,
    ("WC2018", "Australia"):    7.4,
    ("WC2018", "Nigeria"):      7.2,
    ("WC2018", "Morocco"):      7.6,
    ("WC2018", "South Korea"):  7.4,
    ("WC2018", "Iran"):         7.6,
    ("WC2018", "Tunisia"):      7.2,
    ("WC2018", "Costa Rica"):   7.8,
    ("WC2018", "Panama"):       6.8,
    ("WC2018", "Egypt"):        7.4,
    ("WC2018", "Saudi Arabia"): 6.8,
}

# Form at tournament time (last 10 games win ratio, approximate)
HIST_FORM = {
    ("WC2022", "France"):       0.75,
    ("WC2022", "Brazil"):       0.70,
    ("WC2022", "Argentina"):    0.80,  # strong WC qualifier run
    ("WC2022", "England"):      0.65,
    ("WC2022", "Spain"):        0.70,
    ("WC2022", "Portugal"):     0.68,
    ("WC2022", "Netherlands"):  0.68,
    ("WC2022", "Belgium"):      0.60,
    ("WC2022", "Germany"):      0.62,
    ("WC2022", "Croatia"):      0.60,
    ("WC2022", "Morocco"):      0.65,
    ("WC2022", "Switzerland"):  0.64,
    ("WC2022", "Senegal"):      0.65,
    ("WC2022", "Japan"):        0.65,
    ("WC2022", "South Korea"):  0.58,
    ("WC2022", "USA"):          0.62,
    ("WC2022", "Denmark"):      0.65,
    ("WC2022", "Poland"):       0.60,
    ("WC2022", "Uruguay"):      0.62,
    ("WC2022", "Australia"):    0.55,
    ("WC2022", "Ecuador"):      0.58,
    ("WC2022", "Canada"):       0.60,
    ("WC2022", "Serbia"):       0.60,
    ("WC2022", "Tunisia"):      0.54,
    ("WC2022", "Ghana"):        0.50,
    ("WC2022", "Cameroon"):     0.55,
    ("WC2022", "Iran"):         0.58,
    ("WC2022", "Mexico"):       0.60,
    ("WC2022", "Wales"):        0.55,
    ("WC2022", "Saudi Arabia"): 0.52,
    ("WC2022", "Costa Rica"):   0.55,
    ("WC2022", "Qatar"):        0.40,
    ("WC2018", "Germany"):      0.75,  # defending champion but struggling
    ("WC2018", "France"):       0.72,
    ("WC2018", "Brazil"):       0.72,
    ("WC2018", "Argentina"):    0.65,
    ("WC2018", "Spain"):        0.68,
    ("WC2018", "Belgium"):      0.72,
    ("WC2018", "Portugal"):     0.68,
    ("WC2018", "Croatia"):      0.65,
    ("WC2018", "England"):      0.68,
    ("WC2018", "Uruguay"):      0.65,
    ("WC2018", "Colombia"):     0.65,
    ("WC2018", "Switzerland"):  0.65,
    ("WC2018", "Denmark"):      0.62,
    ("WC2018", "Mexico"):       0.62,
    ("WC2018", "Sweden"):       0.62,
    ("WC2018", "Japan"):        0.60,
    ("WC2018", "Senegal"):      0.60,
    ("WC2018", "Russia"):       0.55,
    ("WC2018", "Poland"):       0.62,
    ("WC2018", "Iceland"):      0.58,
    ("WC2018", "Nigeria"):      0.55,
    ("WC2018", "Australia"):    0.52,
    ("WC2018", "Morocco"):      0.58,
    ("WC2018", "Iran"):         0.58,
    ("WC2018", "Peru"):         0.62,
    ("WC2018", "South Korea"):  0.55,
    ("WC2018", "Serbia"):       0.58,
    ("WC2018", "Tunisia"):      0.52,
    ("WC2018", "Costa Rica"):   0.55,
    ("WC2018", "Panama"):       0.48,
    ("WC2018", "Egypt"):        0.55,
    ("WC2018", "Saudi Arabia"): 0.48,
}

# Average age at tournament time
HIST_AGE = {
    ("WC2022", "France"): 26.0, ("WC2022", "Brazil"): 25.5,
    ("WC2022", "Argentina"): 27.2, ("WC2022", "England"): 26.2,
    ("WC2022", "Spain"): 25.5, ("WC2022", "Portugal"): 28.1,
    ("WC2022", "Netherlands"): 26.3, ("WC2022", "Belgium"): 29.0,
    ("WC2022", "Germany"): 25.8, ("WC2022", "Croatia"): 30.1,
    ("WC2022", "Morocco"): 27.2, ("WC2022", "Switzerland"): 27.8,
    ("WC2022", "Senegal"): 26.8, ("WC2022", "Japan"): 25.8,
    ("WC2022", "South Korea"): 27.2, ("WC2022", "USA"): 24.8,
    ("WC2022", "Denmark"): 25.5, ("WC2022", "Poland"): 27.8,
    ("WC2022", "Uruguay"): 27.5, ("WC2022", "Australia"): 27.2,
    ("WC2022", "Ecuador"): 24.8, ("WC2022", "Canada"): 25.5,
    ("WC2022", "Serbia"): 26.8, ("WC2022", "Tunisia"): 27.5,
    ("WC2022", "Ghana"): 25.5, ("WC2022", "Cameroon"): 27.0,
    ("WC2022", "Iran"): 28.0, ("WC2022", "Mexico"): 26.5,
    ("WC2022", "Wales"): 27.8, ("WC2022", "Saudi Arabia"): 26.5,
    ("WC2022", "Costa Rica"): 28.5, ("WC2022", "Qatar"): 26.0,
    ("WC2018", "Germany"): 26.0, ("WC2018", "France"): 25.8,
    ("WC2018", "Brazil"): 25.5, ("WC2018", "Argentina"): 28.5,
    ("WC2018", "Spain"): 28.0, ("WC2018", "Belgium"): 27.5,
    ("WC2018", "Portugal"): 28.5, ("WC2018", "Croatia"): 29.0,
    ("WC2018", "England"): 26.0, ("WC2018", "Uruguay"): 28.5,
    ("WC2018", "Colombia"): 27.0, ("WC2018", "Switzerland"): 27.5,
    ("WC2018", "Denmark"): 26.5, ("WC2018", "Mexico"): 27.0,
    ("WC2018", "Sweden"): 28.0, ("WC2018", "Japan"): 25.0,
    ("WC2018", "Senegal"): 26.5, ("WC2018", "Russia"): 27.8,
    ("WC2018", "Poland"): 27.5, ("WC2018", "Iceland"): 27.8,
    ("WC2018", "Nigeria"): 24.5, ("WC2018", "Australia"): 27.5,
    ("WC2018", "Morocco"): 26.0, ("WC2018", "Iran"): 28.5,
    ("WC2018", "Peru"): 27.5, ("WC2018", "South Korea"): 27.2,
    ("WC2018", "Serbia"): 27.5, ("WC2018", "Tunisia"): 27.5,
    ("WC2018", "Costa Rica"): 28.5, ("WC2018", "Panama"): 28.0,
    ("WC2018", "Egypt"): 27.0, ("WC2018", "Saudi Arabia"): 26.0,
}

# ============================================================
# HEAD-TO-HEAD ADVANTAGE MAP (stable historical data)
# ============================================================
H2H_ADVANTAGE = {
    ("Argentina", "Brazil"):        0.05,
    ("Germany", "Netherlands"):     0.10,
    ("Spain", "Portugal"):          0.08,
    ("England", "Germany"):        -0.05,
    ("Uruguay", "Argentina"):      -0.05,
    ("France", "England"):          0.08,
    ("Mexico", "United States"):    0.12,
    ("Germany", "England"):         0.10,
    ("Brazil", "Argentina"):       -0.05,
    ("Morocco", "Algeria"):         0.05,
    ("South Korea", "Japan"):       0.03,
    ("Iran", "Saudi Arabia"):       0.05,
    ("Egypt", "Algeria"):           0.04,
    ("Netherlands", "Germany"):    -0.10,
    ("Croatia", "Bosnia and Herzegovina"): 0.15,
    ("Spain", "Netherlands"):       0.10,
    ("Italy", "Germany"):           0.05,
    ("Italy", "England"):           0.05,
    ("France", "Germany"):          0.08,
    ("Brazil", "Germany"):         -0.05,
    ("Argentina", "Croatia"):      -0.03,
    ("Spain", "France"):            0.05,
    ("Colombia", "Uruguay"):        0.03,
    ("Belgium", "France"):         -0.05,
    ("Netherlands", "England"):     0.03,
}

def get_h2h(team1, team2):
    if (team1, team2) in H2H_ADVANTAGE:
        return H2H_ADVANTAGE[(team1, team2)]
    if (team2, team1) in H2H_ADVANTAGE:
        return -H2H_ADVANTAGE[(team2, team1)]
    return 0.0

# ============================================================
# VENUE TEMPERATURES 2026
# ============================================================
GROUP_TEMPS = {
    "A": 28, "B": 22, "C": 27, "D": 25,
    "E": 24, "F": 22, "G": 28, "H": 25,
    "I": 26, "J": 24, "K": 28, "L": 23,
}

# ============================================================
# FEATURE ENGINEERING
# ============================================================

def get_team_data(team, tournament=None):
    """Get team stats — historical if tournament given, else 2026."""
    if tournament is None:
        return TEAMS.get(team)
    # Historical lookup
    elo = HIST_ELO.get((tournament, team))
    meta = TEAM_META.get(team, {})
    gk = HIST_GK.get((tournament, team), 7.5)
    form = HIST_FORM.get((tournament, team), 0.57)
    age = HIST_AGE.get((tournament, team), 27.0)
    if elo is None:
        return None
    return {
        "elo": elo,
        "height": meta.get("height", 181.5),
        "age": age,
        "gk": gk,
        "form": form,
        "home_temp": meta.get("home_temp", 20),
    }

def elo_win_prob(elo_a, elo_b):
    return 1.0 / (1.0 + 10 ** ((elo_b - elo_a) / 400.0))

def compute_heat_penalty(team_data, venue_temp):
    """Older teams in unfamiliar heat perform worse."""
    home_temp = team_data.get("home_temp", 20)
    temp_diff = venue_temp - home_temp
    if temp_diff <= 0:
        return 0.0
    age = team_data.get("age", 27.0)
    age_factor = max(0, (age - 25.5) / 4.0)
    return min(temp_diff * 0.004 * (1 + age_factor), 0.20)

def compute_features(t1_data, t2_data, venue_temp, h2h_adv):
    """
    7-dimensional feature vector from team1's perspective.
    All features are continuous; underdog is a soft boost value.
    """
    elo1, elo2 = t1_data["elo"], t2_data["elo"]
    avg_height = 181.5

    elo_diff      = elo1 - elo2
    height_diff   = t1_data["height"] - t2_data["height"]
    heat1         = compute_heat_penalty(t1_data, venue_temp)
    heat2         = compute_heat_penalty(t2_data, venue_temp)
    heat_adv      = heat2 - heat1           # positive = opponent suffers more
    h2h           = h2h_adv
    opp_gk        = t2_data["gk"]           # higher = harder to score against
    form_diff     = t1_data["form"] - t2_data["form"]
    p_win         = elo_win_prob(elo1, elo2)
    underdog      = max(0.0, (0.25 - p_win) / 0.25) if p_win < 0.25 else 0.0

    return [elo_diff, height_diff, heat_adv, h2h, opp_gk, form_diff, underdog]

# ============================================================
# HISTORICAL MATCH RESULTS FOR TRAINING
# Format: (tournament, team1, team2, goals1, goals2, venue_temp)
# ============================================================
HIST_MATCHES = [
    # ── WC 2022 Qatar (all ~28°C) ───────────────────────────
    # Group A
    ("WC2022", "Ecuador",    "Qatar",        2, 0, 28),
    ("WC2022", "Netherlands","Senegal",       2, 0, 28),
    ("WC2022", "Qatar",      "Senegal",       1, 3, 28),
    ("WC2022", "Netherlands","Ecuador",       1, 1, 28),
    ("WC2022", "Ecuador",    "Senegal",       1, 2, 28),
    ("WC2022", "Netherlands","Qatar",         2, 0, 28),
    # Group B
    ("WC2022", "England",    "Iran",          6, 2, 28),
    ("WC2022", "USA",        "Wales",         1, 1, 28),
    ("WC2022", "Wales",      "Iran",          0, 2, 28),
    ("WC2022", "England",    "USA",           0, 0, 28),
    ("WC2022", "England",    "Wales",         3, 0, 28),
    ("WC2022", "Iran",       "USA",           0, 1, 28),
    # Group C
    ("WC2022", "Argentina",  "Saudi Arabia",  1, 2, 28),
    ("WC2022", "Poland",     "Mexico",        0, 0, 28),
    ("WC2022", "Poland",     "Saudi Arabia",  2, 0, 28),
    ("WC2022", "Argentina",  "Mexico",        2, 0, 28),
    ("WC2022", "Poland",     "Argentina",     0, 2, 28),
    ("WC2022", "Saudi Arabia","Mexico",       1, 2, 28),
    # Group D
    ("WC2022", "France",     "Australia",     4, 1, 28),
    ("WC2022", "Denmark",    "Tunisia",       0, 0, 28),
    ("WC2022", "France",     "Denmark",       2, 1, 28),
    ("WC2022", "Australia",  "Tunisia",       1, 0, 28),
    ("WC2022", "Tunisia",    "France",        1, 0, 28),
    ("WC2022", "Australia",  "Denmark",       1, 0, 28),
    # Group E
    ("WC2022", "Spain",      "Costa Rica",    7, 0, 28),
    ("WC2022", "Germany",    "Japan",         1, 2, 28),
    ("WC2022", "Spain",      "Germany",       1, 1, 28),
    ("WC2022", "Japan",      "Costa Rica",    0, 1, 28),
    ("WC2022", "Spain",      "Japan",         1, 2, 28),
    ("WC2022", "Germany",    "Costa Rica",    4, 2, 28),
    # Group F
    ("WC2022", "Belgium",    "Canada",        1, 0, 28),
    ("WC2022", "Morocco",    "Croatia",       0, 0, 28),
    ("WC2022", "Belgium",    "Morocco",       0, 2, 28),
    ("WC2022", "Croatia",    "Canada",        4, 1, 28),
    ("WC2022", "Croatia",    "Belgium",       0, 0, 28),
    ("WC2022", "Morocco",    "Canada",        2, 1, 28),
    # Group G
    ("WC2022", "Brazil",     "Serbia",        2, 0, 28),
    ("WC2022", "Switzerland","Cameroon",      1, 0, 28),
    ("WC2022", "Brazil",     "Switzerland",   1, 0, 28),
    ("WC2022", "Cameroon",   "Serbia",        3, 3, 28),
    ("WC2022", "Brazil",     "Cameroon",      0, 1, 28),
    ("WC2022", "Serbia",     "Switzerland",   2, 3, 28),
    # Group H
    ("WC2022", "Portugal",   "Ghana",         3, 2, 28),
    ("WC2022", "South Korea","Uruguay",       0, 0, 28),
    ("WC2022", "Portugal",   "Uruguay",       2, 0, 28),
    ("WC2022", "South Korea","Ghana",         2, 3, 28),
    ("WC2022", "Portugal",   "South Korea",   1, 2, 28),
    ("WC2022", "Uruguay",    "Ghana",         2, 0, 28),

    # ── WC 2018 Russia (all ~20°C) ──────────────────────────
    # Group A
    ("WC2018", "Russia",     "Saudi Arabia",  5, 0, 20),
    ("WC2018", "Egypt",      "Uruguay",       0, 1, 20),
    ("WC2018", "Russia",     "Egypt",         3, 0, 20),
    ("WC2018", "Uruguay",    "Saudi Arabia",  1, 0, 20),
    ("WC2018", "Saudi Arabia","Egypt",        2, 1, 20),
    ("WC2018", "Uruguay",    "Russia",        3, 0, 20),
    # Group B
    ("WC2018", "Morocco",    "Iran",          0, 1, 20),
    ("WC2018", "Portugal",   "Spain",         3, 3, 20),
    ("WC2018", "Portugal",   "Morocco",       1, 0, 20),
    ("WC2018", "Iran",       "Spain",         0, 1, 20),
    ("WC2018", "Iran",       "Portugal",      1, 1, 20),
    ("WC2018", "Spain",      "Morocco",       2, 2, 20),
    # Group C
    ("WC2018", "France",     "Australia",     2, 1, 20),
    ("WC2018", "Peru",       "Denmark",       0, 1, 20),
    ("WC2018", "France",     "Peru",          1, 0, 20),
    ("WC2018", "Denmark",    "Australia",     1, 1, 20),
    ("WC2018", "France",     "Denmark",       0, 0, 20),
    ("WC2018", "Australia",  "Peru",          0, 2, 20),
    # Group D
    ("WC2018", "Argentina",  "Iceland",       1, 1, 20),
    ("WC2018", "Croatia",    "Nigeria",       2, 0, 20),
    ("WC2018", "Argentina",  "Croatia",       0, 3, 20),
    ("WC2018", "Nigeria",    "Iceland",       2, 0, 20),
    ("WC2018", "Iceland",    "Croatia",       1, 2, 20),
    ("WC2018", "Argentina",  "Nigeria",       2, 1, 20),
    # Group E
    ("WC2018", "Brazil",     "Switzerland",   1, 1, 20),
    ("WC2018", "Serbia",     "Costa Rica",    1, 0, 20),
    ("WC2018", "Brazil",     "Costa Rica",    2, 0, 20),
    ("WC2018", "Serbia",     "Switzerland",   1, 2, 20),
    ("WC2018", "Brazil",     "Serbia",        2, 0, 20),
    ("WC2018", "Switzerland","Costa Rica",    2, 2, 20),
    # Group F
    ("WC2018", "Germany",    "Mexico",        0, 1, 20),
    ("WC2018", "Sweden",     "South Korea",   1, 0, 20),
    ("WC2018", "Germany",    "Sweden",        2, 1, 20),
    ("WC2018", "Mexico",     "South Korea",   2, 1, 20),
    ("WC2018", "South Korea","Germany",       2, 0, 20),
    ("WC2018", "Mexico",     "Sweden",        0, 3, 20),
    # Group G
    ("WC2018", "Belgium",    "Panama",        3, 0, 20),
    ("WC2018", "England",    "Tunisia",       2, 1, 20),
    ("WC2018", "Belgium",    "Tunisia",       5, 2, 20),
    ("WC2018", "England",    "Panama",        6, 1, 20),
    ("WC2018", "England",    "Belgium",       0, 1, 20),
    ("WC2018", "Tunisia",    "Panama",        2, 2, 20),
    # Group H
    ("WC2018", "Colombia",   "Japan",         1, 2, 20),
    ("WC2018", "Poland",     "Senegal",       1, 2, 20),
    ("WC2018", "Japan",      "Senegal",       2, 2, 20),
    ("WC2018", "Poland",     "Colombia",      0, 3, 20),
    ("WC2018", "Japan",      "Poland",        0, 1, 20),
    ("WC2018", "Senegal",    "Colombia",      0, 1, 20),

    # ── Euro 2020/2021 (June-July 2021, avg 20°C) ───────────
    # Group A
    ("EURO2020", "Turkey",      "Italy",          0, 3, 20),
    ("EURO2020", "Wales",       "Switzerland",    1, 1, 20),
    ("EURO2020", "Turkey",      "Wales",          0, 2, 20),
    ("EURO2020", "Italy",       "Switzerland",    3, 0, 20),
    ("EURO2020", "Switzerland", "Turkey",         3, 1, 20),
    ("EURO2020", "Italy",       "Wales",          1, 0, 20),
    # Group B
    ("EURO2020", "Denmark",     "Finland",        0, 1, 20),
    ("EURO2020", "Belgium",     "Russia",         3, 0, 20),
    ("EURO2020", "Finland",     "Russia",         0, 1, 20),
    ("EURO2020", "Denmark",     "Belgium",        1, 2, 20),
    ("EURO2020", "Russia",      "Denmark",        1, 4, 20),
    ("EURO2020", "Finland",     "Belgium",        0, 2, 20),
    # Group C
    ("EURO2020", "Netherlands", "Ukraine",        3, 2, 20),
    ("EURO2020", "Austria",     "North Macedonia",3, 1, 20),
    ("EURO2020", "Ukraine",     "North Macedonia",2, 1, 20),
    ("EURO2020", "Netherlands", "Austria",        2, 0, 20),
    ("EURO2020", "North Macedonia","Netherlands", 0, 3, 20),
    ("EURO2020", "Ukraine",     "Austria",        0, 1, 20),
    # Group D
    ("EURO2020", "England",     "Croatia",        1, 0, 20),
    ("EURO2020", "Scotland",    "Czech Republic", 0, 2, 20),
    ("EURO2020", "Croatia",     "Czech Republic", 1, 1, 20),
    ("EURO2020", "England",     "Scotland",       0, 0, 20),
    ("EURO2020", "Czech Republic","England",      0, 1, 20),
    ("EURO2020", "Croatia",     "Scotland",       3, 1, 20),
    # Group E
    ("EURO2020", "Spain",       "Sweden",         0, 0, 20),
    ("EURO2020", "Poland",      "Slovakia",       1, 2, 20),
    ("EURO2020", "Sweden",      "Slovakia",       1, 0, 20),
    ("EURO2020", "Spain",       "Poland",         1, 1, 20),
    ("EURO2020", "Sweden",      "Poland",         3, 2, 20),
    ("EURO2020", "Slovakia",    "Spain",          0, 5, 20),
    # Group F
    ("EURO2020", "Hungary",     "Portugal",       0, 3, 20),
    ("EURO2020", "France",      "Germany",        1, 0, 20),
    ("EURO2020", "Hungary",     "France",         1, 1, 20),
    ("EURO2020", "Portugal",    "Germany",        2, 4, 20),
    ("EURO2020", "Germany",     "Hungary",        2, 2, 20),
    ("EURO2020", "Portugal",    "France",         2, 2, 20),

    # ── Euro 2024 Germany (June 2024, avg 22°C) ─────────────
    # Group A
    ("EURO2024", "Germany",     "Scotland",       5, 1, 22),
    ("EURO2024", "Hungary",     "Switzerland",    1, 3, 22),
    ("EURO2024", "Germany",     "Hungary",        2, 0, 22),
    ("EURO2024", "Scotland",    "Switzerland",    1, 1, 22),
    ("EURO2024", "Switzerland", "Germany",        1, 1, 22),
    ("EURO2024", "Scotland",    "Hungary",        0, 1, 22),
    # Group B
    ("EURO2024", "Spain",       "Croatia",        3, 0, 22),
    ("EURO2024", "Italy",       "Albania",        2, 1, 22),
    ("EURO2024", "Spain",       "Italy",          1, 0, 22),
    ("EURO2024", "Croatia",     "Albania",        2, 2, 22),
    ("EURO2024", "Albania",     "Spain",          0, 1, 22),
    ("EURO2024", "Croatia",     "Italy",          1, 1, 22),
    # Group C
    ("EURO2024", "Slovenia",    "Denmark",        1, 1, 22),
    ("EURO2024", "Serbia",      "England",        0, 1, 22),
    ("EURO2024", "Slovenia",    "Serbia",         1, 1, 22),
    ("EURO2024", "Denmark",     "England",        1, 1, 22),
    ("EURO2024", "England",     "Slovenia",       0, 0, 22),
    ("EURO2024", "Denmark",     "Serbia",         0, 0, 22),
    # Group D
    ("EURO2024", "Poland",      "Netherlands",    1, 2, 22),
    ("EURO2024", "Austria",     "France",         0, 1, 22),
    ("EURO2024", "Poland",      "Austria",        1, 3, 22),
    ("EURO2024", "Netherlands", "France",         0, 0, 22),
    ("EURO2024", "Netherlands", "Austria",        2, 3, 22),
    ("EURO2024", "France",      "Poland",         1, 1, 22),
    # Group E
    ("EURO2024", "Belgium",     "Slovakia",       0, 1, 22),
    ("EURO2024", "Romania",     "Ukraine",        3, 0, 22),
    ("EURO2024", "Slovakia",    "Ukraine",        1, 2, 22),
    ("EURO2024", "Belgium",     "Romania",        2, 0, 22),
    ("EURO2024", "Ukraine",     "Belgium",        0, 0, 22),
    ("EURO2024", "Slovakia",    "Romania",        1, 1, 22),
    # Group F
    ("EURO2024", "Turkey",      "Georgia",        3, 1, 22),
    ("EURO2024", "Portugal",    "Czech Republic", 2, 1, 22),
    ("EURO2024", "Georgia",     "Czech Republic", 1, 1, 22),
    ("EURO2024", "Turkey",      "Portugal",       0, 3, 22),
    ("EURO2024", "Georgia",     "Portugal",       2, 0, 22),
    ("EURO2024", "Czech Republic","Turkey",       1, 2, 22),

    # ── Copa América 2021 (Brazil, avg 24°C) ────────────────
    ("COPA2021", "Brazil",      "Venezuela",      3, 0, 24),
    ("COPA2021", "Colombia",    "Ecuador",        1, 0, 24),
    ("COPA2021", "Brazil",      "Peru",           4, 0, 24),
    ("COPA2021", "Venezuela",   "Ecuador",        2, 2, 24),
    ("COPA2021", "Brazil",      "Colombia",       2, 1, 24),
    ("COPA2021", "Ecuador",     "Peru",           2, 2, 24),
    ("COPA2021", "Argentina",   "Chile",          1, 1, 24),
    ("COPA2021", "Uruguay",     "Paraguay",       1, 0, 24),
    ("COPA2021", "Argentina",   "Uruguay",        1, 0, 24),
    ("COPA2021", "Chile",       "Paraguay",       0, 0, 24),
    ("COPA2021", "Argentina",   "Paraguay",       1, 0, 24),
    ("COPA2021", "Uruguay",     "Chile",          1, 1, 24),
    ("COPA2021", "Bolivia",     "Paraguay",       2, 1, 24),
    ("COPA2021", "Colombia",    "Venezuela",      3, 0, 24),
    ("COPA2021", "Bolivia",     "Uruguay",        0, 2, 24),
    ("COPA2021", "Ecuador",     "Colombia",       0, 0, 24),

    # ── Copa América 2024 (USA, avg 28°C) ───────────────────
    ("COPA2024", "Argentina",   "Canada",         2, 0, 28),
    ("COPA2024", "Peru",        "Chile",          0, 0, 28),
    ("COPA2024", "Argentina",   "Chile",          1, 0, 28),
    ("COPA2024", "Canada",      "Peru",           1, 0, 28),
    ("COPA2024", "Argentina",   "Peru",           2, 0, 28),
    ("COPA2024", "Canada",      "Chile",          0, 0, 28),
    ("COPA2024", "Colombia",    "Paraguay",       2, 1, 28),
    ("COPA2024", "Brazil",      "Costa Rica",     0, 0, 28),
    ("COPA2024", "Colombia",    "Costa Rica",     3, 0, 28),
    ("COPA2024", "Brazil",      "Paraguay",       4, 1, 28),
    ("COPA2024", "Brazil",      "Colombia",       1, 1, 28),
    ("COPA2024", "Costa Rica",  "Paraguay",       2, 1, 28),
    ("COPA2024", "Uruguay",     "Bolivia",        5, 0, 28),
    ("COPA2024", "Panama",      "USA",            1, 2, 28),
    ("COPA2024", "Uruguay",     "Panama",         3, 1, 28),
    ("COPA2024", "USA",         "Bolivia",        2, 0, 28),
    ("COPA2024", "USA",         "Uruguay",        0, 1, 28),
    ("COPA2024", "Panama",      "Bolivia",        3, 1, 28),
    ("COPA2024", "Ecuador",     "Venezuela",      1, 1, 28),
    ("COPA2024", "Mexico",      "Jamaica",        1, 0, 28),
    ("COPA2024", "Ecuador",     "Jamaica",        3, 1, 28),
    ("COPA2024", "Venezuela",   "Mexico",         1, 0, 28),
    ("COPA2024", "Ecuador",     "Mexico",         0, 0, 28),
    ("COPA2024", "Venezuela",   "Jamaica",        3, 0, 28),

    # ── WC 2014 Brazil (avg 27°C) ───────────────────────────
    # Group A
    ("WC2014", "Brazil",    "Croatia",    3, 1, 27),
    ("WC2014", "Mexico",    "Cameroon",   1, 0, 27),
    ("WC2014", "Brazil",    "Mexico",     0, 0, 27),
    ("WC2014", "Croatia",   "Cameroon",   4, 0, 27),
    ("WC2014", "Croatia",   "Mexico",     1, 3, 27),
    ("WC2014", "Brazil",    "Cameroon",   4, 1, 27),
    # Group B
    ("WC2014", "Netherlands","Spain",     5, 1, 27),
    ("WC2014", "Chile",     "Australia",  3, 1, 27),
    ("WC2014", "Spain",     "Chile",      0, 2, 27),
    ("WC2014", "Netherlands","Australia", 3, 2, 27),
    ("WC2014", "Netherlands","Chile",     2, 0, 27),
    ("WC2014", "Spain",     "Australia",  3, 0, 27),
    # Group C
    ("WC2014", "Colombia",  "Greece",     3, 0, 27),
    ("WC2014", "Ivory Coast","Japan",     2, 1, 27),
    ("WC2014", "Colombia",  "Ivory Coast",2, 1, 27),
    ("WC2014", "Japan",     "Greece",     0, 0, 27),
    ("WC2014", "Colombia",  "Japan",      4, 1, 27),
    ("WC2014", "Ivory Coast","Greece",    1, 2, 27),
    # Group D
    ("WC2014", "Uruguay",   "Costa Rica", 1, 3, 27),
    ("WC2014", "England",   "Italy",      1, 2, 27),
    ("WC2014", "Uruguay",   "England",    2, 1, 27),
    ("WC2014", "Italy",     "Costa Rica", 0, 1, 27),
    ("WC2014", "Costa Rica","England",    0, 0, 27),
    ("WC2014", "Italy",     "Uruguay",    0, 1, 27),
    # Group E
    ("WC2014", "Switzerland","Ecuador",   2, 1, 27),
    ("WC2014", "France",    "Honduras",   3, 0, 27),
    ("WC2014", "France",    "Switzerland",5, 2, 27),
    ("WC2014", "Honduras",  "Ecuador",    1, 2, 27),
    ("WC2014", "France",    "Ecuador",    0, 0, 27),
    ("WC2014", "Switzerland","Honduras",  3, 0, 27),
    # Group F
    ("WC2014", "Argentina", "Bosnia",     2, 1, 27),
    ("WC2014", "Nigeria",   "Iran",       1, 0, 27),
    ("WC2014", "Argentina", "Iran",       1, 0, 27),
    ("WC2014", "Nigeria",   "Bosnia",     1, 0, 27),
    ("WC2014", "Bosnia",    "Iran",       3, 1, 27),
    ("WC2014", "Argentina", "Nigeria",    3, 2, 27),
    # Group G
    ("WC2014", "Germany",   "Portugal",   4, 0, 27),
    ("WC2014", "Ghana",     "USA",        1, 2, 27),
    ("WC2014", "Germany",   "Ghana",      2, 2, 27),
    ("WC2014", "USA",       "Portugal",   2, 2, 27),
    ("WC2014", "USA",       "Germany",    0, 1, 27),
    ("WC2014", "Portugal",  "Ghana",      2, 1, 27),
    # Group H
    ("WC2014", "Belgium",   "Algeria",    2, 1, 27),
    ("WC2014", "Russia",    "South Korea",1, 1, 27),
    ("WC2014", "Belgium",   "Russia",     1, 0, 27),
    ("WC2014", "Algeria",   "South Korea",4, 2, 27),
    ("WC2014", "Algeria",   "Russia",     1, 1, 27),
    ("WC2014", "South Korea","Belgium",   0, 1, 27),

    # ── AFCON 2021 (Cameroon, Jan-Feb 2022, avg 28°C) ───────
    ("AFCON2021", "Cameroon",  "Burkina Faso",  2, 1, 28),
    ("AFCON2021", "Ethiopia",  "Cape Verde",    0, 1, 28),
    ("AFCON2021", "Cameroon",  "Ethiopia",      4, 1, 28),
    ("AFCON2021", "Burkina Faso","Cape Verde",  1, 1, 28),
    ("AFCON2021", "Cameroon",  "Cape Verde",    4, 1, 28),
    ("AFCON2021", "Burkina Faso","Ethiopia",    1, 0, 28),
    ("AFCON2021", "Senegal",   "Zimbabwe",      1, 0, 28),
    ("AFCON2021", "Guinea",    "Malawi",        1, 0, 28),
    ("AFCON2021", "Senegal",   "Guinea",        0, 0, 28),
    ("AFCON2021", "Malawi",    "Zimbabwe",      2, 1, 28),
    ("AFCON2021", "Senegal",   "Malawi",        0, 0, 28),
    ("AFCON2021", "Zimbabwe",  "Guinea",        1, 2, 28),
    ("AFCON2021", "Morocco",   "Ghana",         1, 0, 28),
    ("AFCON2021", "Comoros",   "Gabon",         0, 1, 28),
    ("AFCON2021", "Morocco",   "Comoros",       2, 0, 28),
    ("AFCON2021", "Gabon",     "Ghana",         2, 3, 28),
    ("AFCON2021", "Ghana",     "Comoros",       2, 3, 28),
    ("AFCON2021", "Morocco",   "Gabon",         2, 2, 28),
    ("AFCON2021", "Tunisia",   "Mali",          0, 1, 28),
    ("AFCON2021", "Mauritania","Gambia",        0, 1, 28),
    ("AFCON2021", "Tunisia",   "Mauritania",    4, 0, 28),
    ("AFCON2021", "Mali",      "Gambia",        1, 1, 28),
    ("AFCON2021", "Tunisia",   "Gambia",        1, 0, 28),
    ("AFCON2021", "Mali",      "Mauritania",    2, 0, 28),
    ("AFCON2021", "Nigeria",   "Egypt",         1, 0, 28),
    ("AFCON2021", "Sudan",     "Guinea-Bissau", 0, 0, 28),
    ("AFCON2021", "Nigeria",   "Sudan",         3, 1, 28),
    ("AFCON2021", "Guinea-Bissau","Egypt",      0, 1, 28),
    ("AFCON2021", "Nigeria",   "Guinea-Bissau", 2, 0, 28),
    ("AFCON2021", "Egypt",     "Sudan",         1, 0, 28),
    ("AFCON2021", "Algeria",   "Sierra Leone",  0, 1, 28),
    ("AFCON2021", "Ivory Coast","Equatorial Guinea",0,1,28),
    ("AFCON2021", "Algeria",   "Ivory Coast",   1, 3, 28),
    ("AFCON2021", "Sierra Leone","Equatorial Guinea",1,1,28),
    ("AFCON2021", "Algeria",   "Equatorial Guinea",0,1,28),
    ("AFCON2021", "Sierra Leone","Ivory Coast", 0, 2, 28),
]

# ============================================================
# BUILD TRAINING DATA & TRAIN MODEL
# ============================================================

def build_training_data():
    """
    For each historical match, build one feature vector per team-perspective.
    Features from team1's attacking perspective; target = goals scored.
    """
    X, y = [], []
    skipped = 0
    for row in HIST_MATCHES:
        tourn, t1, t2, g1, g2, vtemp = row
        d1 = get_team_data(t1, tourn)
        d2 = get_team_data(t2, tourn)
        if d1 is None or d2 is None:
            skipped += 1
            continue
        h2h = get_h2h(t1, t2)
        X.append(compute_features(d1, d2, vtemp, h2h))
        y.append(g1)
        h2h_rev = get_h2h(t2, t1)
        X.append(compute_features(d2, d1, vtemp, h2h_rev))
        y.append(g2)
    if skipped:
        print(f"[info] Skipped {skipped} historical matches (missing team data)")
    return np.array(X, dtype=float), np.array(y, dtype=float)


def train_model():
    X, y = build_training_data()
    scaler = StandardScaler()
    X_s = scaler.fit_transform(X)
    # Ridge regression treats goals as continuous; alpha controls regularization
    model = Ridge(alpha=5.0)
    model.fit(X_s, y)

    # Cross-validation to report quality
    cv_scores = cross_val_score(model, X_s, y, cv=5, scoring="neg_mean_squared_error")
    rmse = np.sqrt(-cv_scores.mean())
    print(f"[ML] Training samples: {len(y)}  |  CV RMSE: {rmse:.3f} goals")

    labels = ["elo_diff", "height_diff", "heat_adv", "h2h", "opp_gk", "form_diff", "underdog"]
    coefs = model.coef_
    print("[ML] Learned feature weights (scaled):")
    for label, coef in sorted(zip(labels, coefs), key=lambda x: abs(x[1]), reverse=True):
        print(f"     {label:15s}  {coef:+.4f}")

    return model, scaler


# ============================================================
# 2026 MATCH SIMULATION
# ============================================================

CONF_TEMP = {"UEFA": 12, "CONMEBOL": 22, "CAF": 28, "AFC": 25, "CONCACAF": 25, "OFC": 15}


def _enrich(team, data):
    meta = TEAM_META.get(team, {})
    return {
        "elo": data["elo"],
        "height": data.get("height", meta.get("height", 181.5)),
        "age": data["age"],
        "gk": data["gk"],
        "form": data["form"],
        "home_temp": meta.get("home_temp", CONF_TEMP.get(data.get("conf", "UEFA"), 20)),
    }


def predict_xg(team1, team2, match_id, model, scaler):
    """
    Predict expected goals using the trained model + per-match venue temp
    + Big-3 league adjustment (additive, calibrated at 0.15 max effect).
    """
    venue_temp = MATCH_TEMPS.get(match_id, 25)

    e1 = _enrich(team1, TEAMS[team1])
    e2 = _enrich(team2, TEAMS[team2])

    h2h = get_h2h(team1, team2)
    f1 = np.array(compute_features(e1, e2, venue_temp, h2h)).reshape(1, -1)
    f2 = np.array(compute_features(e2, e1, venue_temp, get_h2h(team2, team1))).reshape(1, -1)

    xg1 = float(model.predict(scaler.transform(f1))[0])
    xg2 = float(model.predict(scaler.transform(f2))[0])

    # Big-3 league adjustment: better weekly competition → slightly higher xG
    b1 = BIG3_PLAYERS.get(team1, 0) / SQUAD_SIZE
    b2 = BIG3_PLAYERS.get(team2, 0) / SQUAD_SIZE
    BIG3_WEIGHT = 0.15
    xg1 += BIG3_WEIGHT * (b1 - b2)
    xg2 += BIG3_WEIGHT * (b2 - b1)

    # Set-piece adjustment: each rating point above opponent adds 0.07 xG.
    # Scale: 5.5 = average, 9.5 = elite (England), 4.0 = weak (Japan, Qatar).
    # A 5-point gap (e.g. England vs Japan) → +0.35 xG — matches the ~30%
    # share of goals that typically come from set pieces in international football.
    SP_WEIGHT = 0.07
    sp1 = TEAMS[team1].get("sp", 5.5)
    sp2 = TEAMS[team2].get("sp", 5.5)
    xg1 += SP_WEIGHT * (sp1 - sp2)
    xg2 += SP_WEIGHT * (sp2 - sp1)

    xg1 = max(0.25, xg1)
    xg2 = max(0.25, xg2)
    return xg1, xg2


def predict_score(team1, team2, match_id, model, scaler):
    """Convert expected goals to a predicted scoreline via Poisson rounding."""
    xg1, xg2 = predict_xg(team1, team2, match_id, model, scaler)

    s1 = int(round(xg1))
    s2 = int(round(xg2))

    p_win = elo_win_prob(TEAMS[team1]["elo"], TEAMS[team2]["elo"])

    # Disambiguate ties: strong favourite should not draw every game
    if s1 == s2:
        if p_win > 0.62:
            s1 += 1
        elif p_win < 0.38:
            s2 += 1

    return s1, s2


# ============================================================
# GROUP STAGE SIMULATION
# ============================================================

MATCHES = [
    (1,  "A", "Mexico",               "South Africa"),
    (2,  "A", "Mexico",               "South Korea"),
    (3,  "A", "Mexico",               "Czechia"),
    (4,  "A", "South Africa",         "South Korea"),
    (5,  "A", "South Africa",         "Czechia"),
    (6,  "A", "South Korea",          "Czechia"),
    (7,  "B", "Canada",               "Bosnia and Herzegovina"),
    (8,  "B", "Canada",               "Qatar"),
    (9,  "B", "Canada",               "Switzerland"),
    (10, "B", "Bosnia and Herzegovina","Qatar"),
    (11, "B", "Bosnia and Herzegovina","Switzerland"),
    (12, "B", "Qatar",                "Switzerland"),
    (13, "C", "Brazil",               "Morocco"),
    (14, "C", "Brazil",               "Haiti"),
    (15, "C", "Brazil",               "Scotland"),
    (16, "C", "Morocco",              "Haiti"),
    (17, "C", "Morocco",              "Scotland"),
    (18, "C", "Haiti",                "Scotland"),
    (19, "D", "United States",        "Paraguay"),
    (20, "D", "United States",        "Australia"),
    (21, "D", "United States",        "Turkiye"),
    (22, "D", "Paraguay",             "Australia"),
    (23, "D", "Paraguay",             "Turkiye"),
    (24, "D", "Australia",            "Turkiye"),
    (25, "E", "Germany",              "Curacao"),
    (26, "E", "Germany",              "Ivory Coast"),
    (27, "E", "Germany",              "Ecuador"),
    (28, "E", "Curacao",              "Ivory Coast"),
    (29, "E", "Curacao",              "Ecuador"),
    (30, "E", "Ivory Coast",          "Ecuador"),
    (31, "F", "Netherlands",          "Japan"),
    (32, "F", "Netherlands",          "Sweden"),
    (33, "F", "Netherlands",          "Tunisia"),
    (34, "F", "Japan",                "Sweden"),
    (35, "F", "Japan",                "Tunisia"),
    (36, "F", "Sweden",               "Tunisia"),
    (37, "G", "Belgium",              "Egypt"),
    (38, "G", "Belgium",              "Iran"),
    (39, "G", "Belgium",              "New Zealand"),
    (40, "G", "Egypt",                "Iran"),
    (41, "G", "Egypt",                "New Zealand"),
    (42, "G", "Iran",                 "New Zealand"),
    (43, "H", "Spain",                "Cape Verde"),
    (44, "H", "Spain",                "Saudi Arabia"),
    (45, "H", "Spain",                "Uruguay"),
    (46, "H", "Cape Verde",           "Saudi Arabia"),
    (47, "H", "Cape Verde",           "Uruguay"),
    (48, "H", "Saudi Arabia",         "Uruguay"),
    (49, "I", "France",               "Senegal"),
    (50, "I", "France",               "Iraq"),
    (51, "I", "France",               "Norway"),
    (52, "I", "Senegal",              "Iraq"),
    (53, "I", "Senegal",              "Norway"),
    (54, "I", "Iraq",                 "Norway"),
    (55, "J", "Argentina",            "Algeria"),
    (56, "J", "Argentina",            "Austria"),
    (57, "J", "Argentina",            "Jordan"),
    (58, "J", "Algeria",              "Austria"),
    (59, "J", "Algeria",              "Jordan"),
    (60, "J", "Austria",              "Jordan"),
    (61, "K", "Portugal",             "DR Congo"),
    (62, "K", "Portugal",             "Uzbekistan"),
    (63, "K", "Portugal",             "Colombia"),
    (64, "K", "DR Congo",             "Uzbekistan"),
    (65, "K", "DR Congo",             "Colombia"),
    (66, "K", "Uzbekistan",           "Colombia"),
    (67, "L", "England",              "Croatia"),
    (68, "L", "England",              "Ghana"),
    (69, "L", "England",              "Panama"),
    (70, "L", "Croatia",              "Ghana"),
    (71, "L", "Croatia",              "Panama"),
    (72, "L", "Ghana",                "Panama"),
]


# ============================================================
# PER-MATCH VENUE TEMPERATURES
# Derived from actual FIFA 2026 schedule (NBC Sports) +
# historical June/July averages per city (SofaScore climate data)
# ============================================================
MATCH_TEMPS = {
    # Group A
    1: 25,   # Mexico City
    2: 30,   # Guadalajara
    3: 25,   # Mexico City
    4: 35,   # Monterrey
    5: 32,   # Atlanta
    6: 30,   # Guadalajara
    # Group B
    7: 24,   # Toronto
    8: 19,   # Vancouver
    9: 19,   # Vancouver
    10: 23,  # Seattle
    11: 24,  # Los Angeles
    12: 26,  # San Francisco
    # Group C
    13: 29,  # New York
    14: 31,  # Philadelphia
    15: 32,  # Miami
    16: 32,  # Atlanta
    17: 28,  # Boston
    18: 28,  # Boston
    # Group D
    19: 24,  # Los Angeles
    20: 23,  # Seattle
    21: 24,  # Los Angeles
    22: 26,  # San Francisco
    23: 26,  # San Francisco
    24: 19,  # Vancouver
    # Group E
    25: 35,  # Houston
    26: 24,  # Toronto
    27: 29,  # New York
    28: 31,  # Philadelphia
    29: 31,  # Kansas City
    30: 31,  # Philadelphia
    # Group F
    31: 36,  # Dallas
    32: 35,  # Houston
    33: 31,  # Kansas City
    34: 36,  # Dallas
    35: 35,  # Monterrey
    36: 35,  # Monterrey
    # Group G
    37: 23,  # Seattle
    38: 24,  # Los Angeles
    39: 19,  # Vancouver
    40: 23,  # Seattle
    41: 19,  # Vancouver
    42: 24,  # Los Angeles
    # Group H
    43: 32,  # Atlanta
    44: 32,  # Atlanta
    45: 30,  # Guadalajara
    46: 35,  # Houston
    47: 32,  # Miami
    48: 32,  # Miami
    # Group I
    49: 29,  # New York
    50: 31,  # Philadelphia
    51: 28,  # Boston
    52: 24,  # Toronto
    53: 29,  # New York
    54: 28,  # Boston
    # Group J
    55: 31,  # Kansas City
    56: 36,  # Dallas
    57: 36,  # Dallas
    58: 31,  # Kansas City
    59: 26,  # San Francisco
    60: 26,  # San Francisco
    # Group K
    61: 35,  # Houston
    62: 35,  # Houston
    63: 32,  # Miami
    64: 32,  # Atlanta
    65: 30,  # Guadalajara
    66: 25,  # Mexico City
    # Group L
    67: 36,  # Dallas
    68: 28,  # Boston
    69: 29,  # New York
    70: 31,  # Philadelphia
    71: 24,  # Toronto
    72: 24,  # Toronto
}

# ============================================================
# BIG-3 LEAGUE PLAYERS PER TEAM
# Premier League (from premierleague.com) + La Liga + Bundesliga
# (La Liga/Bundesliga estimated from 2025-26 squad data)
# Normalised as fraction of 26-player squad in predict functions
# ============================================================
BIG3_PLAYERS = {
    # team: PL + LaLiga + Bundesliga count (out of 26-player squad)
    "France":               13,   # 7PL + 3LaLiga + 3BL
    "Argentina":            10,   # 5PL + 3LaLiga + 2BL
    "England":              23,   # 22PL + 0 + 1BL
    "Brazil":               13,   # 8PL + 3LaLiga + 2BL
    "Spain":                23,   # 7PL + 15LaLiga + 1BL
    "Portugal":             12,   # 7PL + 4LaLiga + 1BL
    "Germany":              18,   # 5PL + 1LaLiga + 12BL
    "Netherlands":          21,   # 15PL + 2LaLiga + 4BL
    "Belgium":              12,   # 8PL + 2LaLiga + 2BL
    "Uruguay":               5,   # 3PL + 1LaLiga + 1BL
    "Croatia":               6,   # 4PL + 1LaLiga + 1BL
    "United States":         5,   # 5PL + 0 + 0
    "Colombia":              5,   # 2PL + 2LaLiga + 1BL
    "Switzerland":           7,   # 4PL + 0 + 3BL
    "Mexico":                2,   # 2PL + 0 + 0
    "Senegal":              10,   # 8PL + 1LaLiga + 1BL
    "Japan":                10,   # 3PL + 0 + 7BL
    "Morocco":               7,   # 4PL + 2LaLiga + 1BL
    "Canada":                3,   # 3PL + 0 + 0
    "South Korea":           4,   # 1PL + 0 + 3BL
    "Ecuador":               5,   # 4PL + 1LaLiga + 0
    "Ivory Coast":           6,   # 4PL + 1LaLiga + 1BL
    "Australia":             0,   # 0PL + 0 + 0
    "Sweden":               11,   # 8PL + 0 + 3BL
    "Norway":               10,   # 7PL + 0 + 3BL
    "Algeria":               3,   # 1PL + 1LaLiga + 1BL
    "Austria":               7,   # 2PL + 0 + 5BL
    "Turkiye":               3,   # 2PL + 0 + 1BL
    "Paraguay":              2,   # 2PL + 0 + 0
    "Iran":                  0,   # 0PL + 0 + 0
    "Tunisia":               2,   # 1PL + 1LaLiga + 0
    "Egypt":                 2,   # 2PL + 0 + 0
    "Ghana":                 3,   # 2PL + 0 + 1BL
    "Bosnia and Herzegovina":1,   # 0PL + 0 + 1BL
    "Scotland":              9,   # 9PL + 0 + 0
    "Saudi Arabia":          0,   # 0
    "Panama":                0,   # 0
    "Jordan":                0,   # 0
    "New Zealand":           3,   # 3PL + 0 + 0
    "DR Congo":              7,   # 5PL + 1LaLiga + 1BL
    "Uzbekistan":            1,   # 1PL + 0 + 0
    "Cape Verde":            0,   # 0
    "Czechia":               4,   # 2PL + 0 + 2BL
    "Iraq":                  1,   # 1PL + 0 + 0
    "South Africa":          1,   # 1PL + 0 + 0
    "Curacao":               0,   # 0
    "Qatar":                 0,   # 0
    "Haiti":                 2,   # 2PL + 0 + 0
}

SQUAD_SIZE = 26  # WC2026 squads are 26 players


def compute_standings(results):
    teams = {}
    for _, _, t1, t2, s1, s2 in results:
        for t in [t1, t2]:
            if t not in teams:
                teams[t] = {"pts": 0, "gd": 0, "gf": 0, "ga": 0, "w": 0, "d": 0, "l": 0}
        if s1 > s2:
            teams[t1]["pts"] += 3; teams[t1]["w"] += 1; teams[t2]["l"] += 1
        elif s1 < s2:
            teams[t2]["pts"] += 3; teams[t2]["w"] += 1; teams[t1]["l"] += 1
        else:
            teams[t1]["pts"] += 1; teams[t1]["d"] += 1
            teams[t2]["pts"] += 1; teams[t2]["d"] += 1
        teams[t1]["gf"] += s1; teams[t1]["ga"] += s2; teams[t1]["gd"] += s1 - s2
        teams[t2]["gf"] += s2; teams[t2]["ga"] += s1; teams[t2]["gd"] += s2 - s1
    return sorted(teams.items(), key=lambda x: (-x[1]["pts"], -x[1]["gd"], -x[1]["gf"]))


# ============================================================
# MAIN
# ============================================================

MATCH_DATE = {
    1:"Jun 11", 2:"Jun 18", 3:"Jun 24", 4:"Jun 24", 5:"Jun 18", 6:"Jun 11",
    7:"Jun 12", 8:"Jun 24", 9:"Jun 24", 10:"Jun 24", 11:"Jun 18", 12:"Jun 13",
    13:"Jun 13", 14:"Jun 19", 15:"Jun 24", 16:"Jun 24", 17:"Jun 19", 18:"Jun 13",
    19:"Jun 12", 20:"Jun 19", 21:"Jun 25", 22:"Jun 25", 23:"Jun 19", 24:"Jun 13",
    25:"Jun 14", 26:"Jun 20", 27:"Jun 25", 28:"Jun 25", 29:"Jun 20", 30:"Jun 14",
    31:"Jun 14", 32:"Jun 20", 33:"Jun 25", 34:"Jun 25", 35:"Jun 25", 36:"Jun 20",
    37:"Jun 15", 38:"Jun 21", 39:"Jun 26", 40:"Jun 26", 41:"Jun 21", 42:"Jun 15",
    43:"Jun 15", 44:"Jun 21", 45:"Jun 26", 46:"Jun 26", 47:"Jun 26", 48:"Jun 26",
    49:"Jun 16", 50:"Jun 22", 51:"Jun 26", 52:"Jun 22", 53:"Jun 26", 54:"Jun 26",
    55:"Jun 16", 56:"Jun 22", 57:"Jun 27", 58:"Jun 27", 59:"Jun 22", 60:"Jun 27",
    61:"Jun 17", 62:"Jun 23", 63:"Jun 27", 64:"Jun 27", 65:"Jun 27", 66:"Jun 23",
    67:"Jun 17", 68:"Jun 23", 69:"Jun 27", 70:"Jun 23", 71:"Jun 27", 72:"Jun 27",
}

MATCH_VENUE = {
    1:"Mexico City", 2:"Guadalajara", 3:"Mexico City", 4:"Monterrey",
    5:"Atlanta", 6:"Guadalajara", 7:"Toronto", 8:"Vancouver", 9:"Vancouver",
    10:"Seattle", 11:"Los Angeles", 12:"San Francisco", 13:"New York",
    14:"Philadelphia", 15:"Miami", 16:"Atlanta", 17:"Boston", 18:"Boston",
    19:"Los Angeles", 20:"Seattle", 21:"Los Angeles", 22:"San Francisco",
    23:"San Francisco", 24:"Vancouver", 25:"Houston", 26:"Toronto",
    27:"New York", 28:"Philadelphia", 29:"Kansas City", 30:"Philadelphia",
    31:"Dallas", 32:"Houston", 33:"Kansas City", 34:"Dallas", 35:"Monterrey",
    36:"Monterrey", 37:"Seattle", 38:"Los Angeles", 39:"Vancouver",
    40:"Seattle", 41:"Vancouver", 42:"Los Angeles", 43:"Atlanta",
    44:"Atlanta", 45:"Guadalajara", 46:"Houston", 47:"Miami", 48:"Miami",
    49:"New York", 50:"Philadelphia", 51:"Boston", 52:"Toronto",
    53:"New York", 54:"Boston", 55:"Kansas City", 56:"Dallas", 57:"Dallas",
    58:"Kansas City", 59:"San Francisco", 60:"San Francisco", 61:"Houston",
    62:"Houston", 63:"Miami", 64:"Atlanta", 65:"Guadalajara", 66:"Mexico City",
    67:"Dallas", 68:"Boston", 69:"New York", 70:"Philadelphia",
    71:"Toronto", 72:"Toronto",
}


def export_excel(all_results, group_map, model, scaler, path):
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── colour palette ──────────────────────────────────────
    WIN_FILL   = PatternFill("solid", fgColor="C6EFCE")   # green
    DRAW_FILL  = PatternFill("solid", fgColor="FFEB9C")   # amber
    LOSS_FILL  = PatternFill("solid", fgColor="FFC7CE")   # red
    HDR_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
    GRP_FILL   = PatternFill("solid", fgColor="2E75B6")   # mid blue
    Q_FILL     = PatternFill("solid", fgColor="C6EFCE")   # green → qualified
    Q3_FILL    = PatternFill("solid", fgColor="FFEB9C")   # amber  → maybe
    HDR_FONT   = Font(bold=True, color="FFFFFF")
    GRP_FONT   = Font(bold=True, color="FFFFFF", size=12)
    BOLD       = Font(bold=True)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_col_width(ws, col, width):
        ws.column_dimensions[get_column_letter(col)].width = width

    def hdr_cell(ws, row, col, val, fill=None, font=None, align="center"):
        c = ws.cell(row=row, column=col, value=val)
        if fill: c.fill = fill
        if font: c.font = font
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = border
        return c

    # ── Sheet 1: Match Predictions ───────────────────────────
    ws1 = wb.active
    ws1.title = "Match Predictions"
    ws1.freeze_panes = "A3"
    ws1.row_dimensions[1].height = 20
    ws1.row_dimensions[2].height = 20

    headers = ["#", "Date", "Group", "Team 1", "xG 1", "Score",
               "xG 2", "Team 2", "Result", "Venue", "Temp °C", "Big3 T1", "Big3 T2"]
    for c, h in enumerate(headers, 1):
        hdr_cell(ws1, 1, c, h, fill=HDR_FILL, font=HDR_FONT)

    col_widths = [5, 8, 7, 24, 7, 7, 7, 24, 7, 16, 9, 10, 10]
    for i, w in enumerate(col_widths, 1):
        set_col_width(ws1, i, w)

    sorted_results = sorted(all_results, key=lambda r: r[0])
    for row_i, (mid, grp, t1, t2, s1, s2) in enumerate(sorted_results, 2):
        xg1, xg2 = predict_xg(t1, t2, mid, model, scaler)
        temp = MATCH_TEMPS.get(mid, 25)
        venue = MATCH_VENUE.get(mid, "")
        date = MATCH_DATE.get(mid, "")
        b1 = round(BIG3_PLAYERS.get(t1, 0) / SQUAD_SIZE * 100)
        b2 = round(BIG3_PLAYERS.get(t2, 0) / SQUAD_SIZE * 100)

        if s1 > s2:
            result, fill = f"{t1} wins", WIN_FILL
        elif s1 < s2:
            result, fill = f"{t2} wins", LOSS_FILL
        else:
            result, fill = "Draw", DRAW_FILL

        vals = [mid, date, grp, t1, round(xg1,2), f"{s1} - {s2}",
                round(xg2,2), t2, result, venue, temp, f"{b1}%", f"{b2}%"]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=row_i, column=c, value=v)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if c in (4, 8):
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Sheet 2: Group Standings ─────────────────────────────
    ws2 = wb.create_sheet("Group Standings")
    ws2.freeze_panes = "A2"
    row = 1
    for group_id in sorted(group_map.keys()):
        group_results = [r for r in all_results if r[1] == group_id]
        standings = compute_standings(group_results)

        # Group header
        hdr_cell(ws2, row, 1, f"Group {group_id}", fill=GRP_FILL, font=GRP_FONT)
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1

        col_hdrs = ["Pos", "Team", "P", "W", "D", "L", "GF", "GA", "GD", "Pts"]
        for c, h in enumerate(col_hdrs, 1):
            hdr_cell(ws2, row, c, h, fill=HDR_FILL, font=HDR_FONT)
        row += 1

        for pos, (team, st) in enumerate(standings, 1):
            if pos <= 2:
                row_fill = Q_FILL
            elif pos == 3:
                row_fill = Q3_FILL
            else:
                row_fill = None
            vals = [pos, team, st["w"]+st["d"]+st["l"],
                    st["w"], st["d"], st["l"],
                    st["gf"], st["ga"], st["gd"], st["pts"]]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(row=row, column=c, value=v)
                if row_fill: cell.fill = row_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                if c == 2:
                    cell.alignment = Alignment(horizontal="left")
            row += 1
        row += 1  # blank spacer between groups

    set_col_width(ws2, 1, 5)
    set_col_width(ws2, 2, 26)
    for c in range(3, 11):
        set_col_width(ws2, c, 6)

    # ── Sheet 3: Qualifiers Summary ──────────────────────────
    ws3 = wb.create_sheet("Qualifiers")
    hdr_row = ["Group", "1st Place", "2nd Place", "Best 3rd (candidate)"]
    for c, h in enumerate(hdr_row, 1):
        hdr_cell(ws3, 1, c, h, fill=HDR_FILL, font=HDR_FONT)
    set_col_width(ws3, 1, 8)
    set_col_width(ws3, 2, 28)
    set_col_width(ws3, 3, 28)
    set_col_width(ws3, 4, 28)

    for row_i, group_id in enumerate(sorted(group_map.keys()), 2):
        group_results = [r for r in all_results if r[1] == group_id]
        standings = compute_standings(group_results)
        first  = standings[0][0] if len(standings) > 0 else "?"
        second = standings[1][0] if len(standings) > 1 else "?"
        third  = standings[2][0] if len(standings) > 2 else "?"
        for c, v in enumerate([group_id, first, second, third], 1):
            cell = ws3.cell(row=row_i, column=c, value=v)
            cell.border = border
            fill = Q_FILL if c in (2, 3) else Q3_FILL if c == 4 else None
            if fill: cell.fill = fill
            cell.alignment = Alignment(horizontal="left" if c > 1 else "center")

    # ── Sheet 4: Team Features ───────────────────────────────
    ws4 = wb.create_sheet("Team Features")
    feat_hdrs = ["Team", "ELO", "Height cm", "Avg Age", "GK Rating",
                 "Form %", "SP Rating", "Big3 Players", "Big3 %"]
    for c, h in enumerate(feat_hdrs, 1):
        hdr_cell(ws4, 1, c, h, fill=HDR_FILL, font=HDR_FONT)
    set_col_width(ws4, 1, 28)
    for c in range(2, 10):
        set_col_width(ws4, c, 13)

    for row_i, (team, data) in enumerate(sorted(TEAMS.items()), 2):
        b3 = BIG3_PLAYERS.get(team, 0)
        vals = [team, data["elo"], TEAM_META.get(team, {}).get("height", 181.5),
                data["age"], data["gk"], round(data["form"]*100),
                data.get("sp", 5.5), b3, f"{round(b3/SQUAD_SIZE*100)}%"]
        for c, v in enumerate(vals, 1):
            cell = ws4.cell(row=row_i, column=c, value=v)
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if c == 1 else "center")

    wb.save(path)
    print(f"Excel → {path}")


def main():
    print("=" * 65)
    print("WC2026 GROUP STAGE PREDICTOR — ML Edition")
    print("Training on ~10 years of international football data")
    print("=" * 65)

    model, scaler = train_model()

    all_results = []
    group_map = defaultdict(list)
    for match in MATCHES:
        group_map[match[1]].append(match)

    print("\n" + "=" * 65)

    for group_id in sorted(group_map.keys()):
        results = []
        for match_id, grp, t1, t2 in group_map[group_id]:
            s1, s2 = predict_score(t1, t2, match_id, model, scaler)
            results.append((match_id, grp, t1, t2, s1, s2))
        all_results.extend(results)

        print(f"\n--- Group {group_id} ---")
        for mid, grp, t1, t2, s1, s2 in results:
            xg1, xg2 = predict_xg(t1, t2, mid, model, scaler)
            outcome = "WIN" if s1 > s2 else ("DRAW" if s1 == s2 else "LOSS")
            temp = MATCH_TEMPS.get(mid, 25)
            print(f"  {t1:28s} {s1}-{s2}  {t2:28s}  [{outcome}]  "
                  f"xG:{xg1:.2f}-{xg2:.2f}  {temp}°C {MATCH_VENUE.get(mid,'')}")

        standings = compute_standings(results)
        print(f"  Standings:")
        for i, (team, st) in enumerate(standings):
            q = "Q" if i < 2 else ("3rd" if i == 2 else " ")
            print(f"  {i+1}. {team:28s} {st['pts']}pts  GD:{st['gd']:+d}  [{q}]")

    excel_path = "/Users/melswittenberg/Documents/github/WC26 - prediction model/WC2026_Predictions.xlsx"
    export_excel(all_results, group_map, model, scaler, excel_path)

    # CSV output — matches output_template 1.csv format exactly
    csv_path = "/Users/melswittenberg/Documents/github/WC26 - prediction model/output.csv"
    results_by_id = {mid: (grp, t1, t2, s1, s2)
                     for mid, grp, t1, t2, s1, s2 in all_results}
    with open(csv_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["match_id", "group", "team1", "team2", "score1", "score2"])
        for mid, grp, t1, t2 in MATCHES:
            grp2, t1r, t2r, s1, s2 = results_by_id[mid]
            writer.writerow([mid, grp2, t1r, t2r, s1, s2])
    print(f"CSV   → {csv_path}")

    print("\n--- Predicted Qualifiers (Top 2 per group) ---")
    for group_id in sorted(group_map.keys()):
        group_results = [r for r in all_results if r[1] == group_id]
        standings = compute_standings(group_results)
        q = [standings[0][0], standings[1][0]]
        third = standings[2][0] if len(standings) > 2 else "?"
        print(f"  Group {group_id}: {q[0]} & {q[1]}  (3rd: {third})")


if __name__ == "__main__":
    main()
